# -*- coding: utf-8 -*-
"""
Export laporan keuangan ke file Excel.
Format mirip Finance_Tracker_Perusahaan.xlsx (multi-sheet, formulas, charts).
"""
import os
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from db import Account, Transaction, JournalLine
from accounting import (
    account_balance, all_balances,
    report_income_statement, report_balance_sheet, report_cash_flow,
    report_by_category,
)
from budget import budget_status, progress_bar, fmt_rp as _fmt_rp


# Styles
PRIMARY_FILL  = PatternFill("solid", fgColor="1E3A8A")
HEADER_FILL   = PatternFill("solid", fgColor="2563EB")
SUBHEAD_FILL  = PatternFill("solid", fgColor="DBEAFE")
TOTAL_FILL    = PatternFill("solid", fgColor="FEF3C7")
SECTION_FILL  = PatternFill("solid", fgColor="0F172A")
ZEBRA_FILL    = PatternFill("solid", fgColor="F8FAFC")
PROFIT_FILL   = PatternFill("solid", fgColor="BBF7D0")
LOSS_FILL     = PatternFill("solid", fgColor="FECACA")

HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT   = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
TOTAL_FONT   = Font(name="Calibri", size=11, bold=True)
BODY_FONT    = Font(name="Calibri", size=10)

thin = Side(border_style="thin", color="CBD5E1")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
medium = Side(border_style="medium", color="1E3A8A")
HEAVY_BORDER = Border(left=medium, right=medium, top=medium, bottom=medium)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

CURRENCY_FMT = '"Rp "#,##0;[Red]"Rp ("#,##0")";"Rp "-'
PCT_FMT      = "0.00%"


def _fmt(n):
    if n is None: return 0
    return float(n)


def export_to_excel(session, output_path: str = None) -> str:
    """Generate workbook Excel lengkap. Returns path."""
    if output_path is None:
        out_dir = Path(os.path.dirname(__file__)) / "exports"
        out_dir.mkdir(exist_ok=True)
        output_path = str(out_dir / f"Finance_Report_{datetime.utcnow():%Y%m%d_%H%M}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # hapus sheet default

    # ====== SHEET 1: DASHBOARD ======
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H3")
    ws["B2"] = "FINANCE TRACKER — EXPORT"
    ws["B2"].font = TITLE_FONT
    ws["B2"].fill = PRIMARY_FILL
    ws["B2"].alignment = CENTER

    ws.merge_cells("B4:H4")
    ws["B4"] = f"Export: {datetime.utcnow():%d %B %Y %H:%M} UTC"
    ws["B4"].fill = SUBHEAD_FILL
    ws["B4"].alignment = CENTER

    # KPI cards
    ws["B7"] = "💰 PEMASUKAN"
    ws["C7"] = "💸 PENGELUARAN"
    ws["D7"] = "📊 LABA BERSIH"
    ws["E7"] = "💵 SALDO KAS"
    for col in ["B7","C7","D7","E7"]:
        ws[col].font = HEADER_FONT; ws[col].fill = HEADER_FILL; ws[col].alignment = CENTER; ws[col].border = THIN_BORDER

    r = report_income_statement(session)
    bs = report_balance_sheet(session)
    cf = report_cash_flow(session)
    ws["B8"] = _fmt(r["income"]); ws["B8"].number_format = CURRENCY_FMT
    ws["C8"] = _fmt(r["expense"]); ws["C8"].number_format = CURRENCY_FMT
    ws["D8"] = _fmt(r["net"]); ws["D8"].number_format = CURRENCY_FMT
    ws["E8"] = _fmt(cf["saldo_akhir"]); ws["E8"].number_format = CURRENCY_FMT
    for col in ["B8","C8","D8","E8"]:
        ws[col].font = Font(name="Calibri", size=14, bold=True)
        ws[col].alignment = CENTER; ws[col].border = HEAVY_BORDER

    # Margin
    margin = (r["net"]/r["income"]) if r["income"] else 0
    ws["B9"] = "Margin:"; ws["B9"].font = TOTAL_FONT; ws["B9"].alignment = RIGHT
    ws["C9"] = margin; ws["C9"].number_format = PCT_FMT
    ws["C9"].font = TOTAL_FONT; ws["C9"].alignment = LEFT
    ws.merge_cells("C9:E9")
    ws["B9"].border = THIN_BORDER
    for col in ["C9","D9","E9"]: ws[col].border = THIN_BORDER

    # ====== SHEET 2: JURNAL UMUM ======
    ws = wb.create_sheet("Jurnal Umum")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:I2")
    ws["B2"] = "JURNAL UMUM — DOUBLE ENTRY"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = PRIMARY_FILL; ws["B2"].alignment = CENTER

    headers = ["Tanggal", "No. Jurnal", "Keterangan", "Kode Akun", "Kategori", "Debit", "Kredit"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=2+i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    txns = session.query(Transaction).order_by(Transaction.id).all()
    cur = 5
    for t in txns:
        ws.cell(row=cur, column=2, value=t.txn_date).number_format = "yyyy-mm-dd"
        ws.cell(row=cur, column=3, value=t.jv_number)
        ws.cell(row=cur, column=4, value=t.description)
        # Ambil akun dari line pertama (Debit & Kredit parallel baris)
        d_line = next((l for l in t.lines if (l.debit or 0) > 0), None)
        k_line = next((l for l in t.lines if (l.credit or 0) > 0), None)
        if d_line:
            ws.cell(row=cur, column=5, value=d_line.account_code)
            ws.cell(row=cur, column=7, value=_fmt(d_line.debit))
        if k_line:
            # kalau debit & kredit di row sama, tampilkan di baris yg sama
            if d_line is k_line:
                ws.cell(row=cur, column=8, value=_fmt(k_line.credit))
            else:
                # baris terpisah
                ws.cell(row=cur, column=5, value=f"{d_line.account_code}/{k_line.account_code}")
                ws.cell(row=cur, column=7, value=_fmt(d_line.debit))
                ws.cell(row=cur, column=8, value=_fmt(k_line.credit))
        ws.cell(row=cur, column=6, value=t.category or "")
        for col in range(2, 9):
            cell = ws.cell(row=cur, column=col)
            cell.font = BODY_FONT; cell.border = THIN_BORDER
            cell.alignment = LEFT if col<=6 else RIGHT
            if col in (7,8): cell.number_format = CURRENCY_FMT
            if (cur % 2) == 1: cell.fill = ZEBRA_FILL
        cur += 1
    # Total
    ws.cell(row=cur, column=2, value="TOTAL")
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=6)
    ws.cell(row=cur, column=7, value=f"=SUM(G5:G{cur-1})").number_format = CURRENCY_FMT
    ws.cell(row=cur, column=8, value=f"=SUM(H5:H{cur-1})").number_format = CURRENCY_FMT
    for col in range(2,9):
        cell = ws.cell(row=cur, column=col)
        cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=7 else LEFT

    # Column widths
    widths = {"A":2,"B":12,"C":12,"D":36,"E":12,"F":14,"G":16,"H":16}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    # ====== SHEET 3: BUKU BESAR ======
    ws = wb.create_sheet("Buku Besar")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:I2")
    ws["B2"] = "BUKU BESAR"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = PRIMARY_FILL; ws["B2"].alignment = CENTER

    rows = all_balances(session)
    cur = 4
    sub_h = ["Kode", "Nama Akun", "Tipe", "Normal", "Debit", "Kredit", "Saldo"]
    for i, h in enumerate(sub_h):
        c = ws.cell(row=cur, column=2+i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    cur += 1
    start_r = cur
    for i, (acc, d, k, saldo) in enumerate(rows):
        ws.cell(row=cur, column=2, value=acc.code).font = BODY_FONT
        ws.cell(row=cur, column=3, value=acc.name).font = BODY_FONT
        ws.cell(row=cur, column=4, value=acc.type).font = BODY_FONT
        ws.cell(row=cur, column=5, value=acc.normal_side).font = BODY_FONT
        ws.cell(row=cur, column=6, value=_fmt(d)).number_format = CURRENCY_FMT
        ws.cell(row=cur, column=7, value=_fmt(k)).number_format = CURRENCY_FMT
        ws.cell(row=cur, column=8, value=_fmt(saldo)).number_format = CURRENCY_FMT
        for col in range(2,9):
            cell = ws.cell(row=cur, column=col)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col<=5 else RIGHT
            if i%2==1: cell.fill = ZEBRA_FILL
        cur += 1
    end_r = cur - 1
    ws.cell(row=cur, column=2, value="TOTAL")
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=5)
    ws.cell(row=cur, column=6, value=f"=SUM(F{start_r}:F{end_r})").number_format = CURRENCY_FMT
    ws.cell(row=cur, column=7, value=f"=SUM(G{start_r}:G{end_r})").number_format = CURRENCY_FMT
    ws.cell(row=cur, column=8, value=f"=SUM(H{start_r}:H{end_r})").number_format = CURRENCY_FMT
    for col in range(2,9):
        cell = ws.cell(row=cur, column=col)
        cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=6 else LEFT
    ws.conditional_formatting.add(f"H{start_r}:H{end_r}", CellIsRule(operator="lessThan", formula=["0"], fill=LOSS_FILL))
    ws.conditional_formatting.add(f"H{start_r}:H{end_r}", CellIsRule(operator="greaterThan", formula=["0"], fill=PROFIT_FILL))

    widths = {"A":2,"B":10,"C":28,"D":14,"E":12,"F":18,"G":18,"H":18}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    # ====== SHEET 4: LAPORAN L/R ======
    ws = wb.create_sheet("Laporan L_R")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:D2")
    ws["B2"] = "LAPORAN LABA RUGI"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = PRIMARY_FILL; ws["B2"].alignment = CENTER

    cur = 4
    ws.cell(row=cur, column=2, value="PENDAPATAN").font = HEADER_FONT
    ws.cell(row=cur, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
    ws.cell(row=cur, column=2).alignment = LEFT
    cur += 1
    income_start = cur
    for name, amt in r["income_items"]:
        if amt == 0: continue
        ws.cell(row=cur, column=2, value=f"  {name}").font = BODY_FONT
        ws.cell(row=cur, column=3, value=_fmt(amt)).number_format = CURRENCY_FMT
        for col in (2,3,4):
            cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
            cell.alignment = LEFT if col==2 else RIGHT
        cur += 1
    income_end = cur - 1
    ws.cell(row=cur, column=2, value="Total Pendapatan").font = TOTAL_FONT
    ws.cell(row=cur, column=3, value=f"=SUM(C{income_start}:C{income_end})").number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT
    sub_inc = cur

    cur += 2
    ws.cell(row=cur, column=2, value="BEBAN OPERASIONAL").font = HEADER_FONT
    ws.cell(row=cur, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
    ws.cell(row=cur, column=2).alignment = LEFT
    cur += 1
    exp_start = cur
    for name, amt in r["expense_items"]:
        if amt == 0: continue
        ws.cell(row=cur, column=2, value=f"  {name}").font = BODY_FONT
        ws.cell(row=cur, column=4, value=_fmt(amt)).number_format = CURRENCY_FMT
        for col in (2,3,4):
            cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
            cell.alignment = LEFT if col==2 else RIGHT
        cur += 1
    exp_end = cur - 1
    ws.cell(row=cur, column=2, value="Total Beban").font = TOTAL_FONT
    ws.cell(row=cur, column=4, value=f"=SUM(D{exp_start}:D{exp_end})").number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT
    sub_exp = cur

    cur += 1
    ws.cell(row=cur, column=2, value="LABA (RUGI) BERSIH").font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(row=cur, column=3, value=f"=C{sub_inc}-D{sub_exp}").number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.fill = SECTION_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT

    widths = {"A":2,"B":34,"C":18,"D":18}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    # ====== SHEET 5: NERACA ======
    ws = wb.create_sheet("Neraca")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:D2")
    ws["B2"] = "NERACA"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = PRIMARY_FILL; ws["B2"].alignment = CENTER

    cur = 4
    # Aset
    ws.cell(row=cur, column=2, value="ASET").font = HEADER_FONT
    ws.cell(row=cur, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
    cur += 1
    for acc, d, k, saldo in rows:
        if acc.type == "Aset" and saldo != 0:
            ws.cell(row=cur, column=2, value=f"  {acc.name}").font = BODY_FONT
            ws.cell(row=cur, column=3, value=_fmt(saldo)).number_format = CURRENCY_FMT
            for col in (2,3,4):
                cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
                cell.alignment = LEFT if col==2 else RIGHT
            cur += 1
    ws.cell(row=cur, column=2, value="TOTAL ASET").font = TOTAL_FONT
    ws.cell(row=cur, column=3, value=_fmt(bs["aset"])).number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT

    cur += 2
    # Liabilitas
    ws.cell(row=cur, column=2, value="LIABILITAS").font = HEADER_FONT
    ws.cell(row=cur, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
    cur += 1
    for acc, d, k, saldo in rows:
        if acc.type == "Liabilitas" and saldo != 0:
            ws.cell(row=cur, column=2, value=f"  {acc.name}").font = BODY_FONT
            ws.cell(row=cur, column=3, value=_fmt(saldo)).number_format = CURRENCY_FMT
            for col in (2,3,4):
                cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
                cell.alignment = LEFT if col==2 else RIGHT
            cur += 1
    if bs["liabilitas"] == 0:
        ws.cell(row=cur, column=2, value="  (tidak ada liabilitas)").font = BODY_FONT
        cur += 1
    ws.cell(row=cur, column=2, value="TOTAL LIABILITAS").font = TOTAL_FONT
    ws.cell(row=cur, column=3, value=_fmt(bs["liabilitas"])).number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT

    cur += 2
    # Ekuitas
    ws.cell(row=cur, column=2, value="EKUITAS").font = HEADER_FONT
    ws.cell(row=cur, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
    cur += 1
    for acc, d, k, saldo in rows:
        if acc.type == "Ekuitas" and saldo != 0:
            ws.cell(row=cur, column=2, value=f"  {acc.name}").font = BODY_FONT
            ws.cell(row=cur, column=3, value=_fmt(saldo)).number_format = CURRENCY_FMT
            for col in (2,3,4):
                cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
                cell.alignment = LEFT if col==2 else RIGHT
            cur += 1
    ws.cell(row=cur, column=2, value="  Laba Tahun Berjalan").font = BODY_FONT
    ws.cell(row=cur, column=3, value=_fmt(bs["laba"])).number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.border = THIN_BORDER
        cell.alignment = LEFT if col==2 else RIGHT
    cur += 1
    ws.cell(row=cur, column=2, value="TOTAL EKUITAS").font = TOTAL_FONT
    ws.cell(row=cur, column=3, value=_fmt(bs["total_ekuitas"])).number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT

    cur += 1
    ws.cell(row=cur, column=2, value="TOTAL LIAB + EKUITAS").font = TOTAL_FONT
    ws.cell(row=cur, column=3, value=_fmt(bs["liabilitas"] + bs["total_ekuitas"])).number_format = CURRENCY_FMT
    for col in (2,3,4):
        cell = ws.cell(row=cur, column=col); cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = HEAVY_BORDER
        cell.alignment = RIGHT if col>=3 else LEFT

    widths = {"A":2,"B":34,"C":20,"D":4}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    # ====== SHEET 6: KATEGORI ======
    ws = wb.create_sheet("Kategori")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:F2")
    ws["B2"] = "PENGELUARAN PER KATEGORI"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = PRIMARY_FILL; ws["B2"].alignment = CENTER

    cat_totals = report_by_category(session)
    cur = 4
    headers = ["Kategori", "Target %", "Target (Rp)", "Realisasi (Rp)", "Selisih"]
    for i, h in enumerate(headers):
        c = ws.cell(row=cur, column=2+i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    cur += 1
    # Ambil budgets
    from budget import all_budgets
    budgets = all_budgets(session)
    from accounts_seed import CATEGORY_LIST
    cat_label = {"operasional":"Operasional","gaji":"Gaji","pemasaran":"Pemasaran",
                 "peralatan":"Peralatan","utilitas":"Utilitas","transport":"Transport"}
    for cat in CATEGORY_LIST:
        real = cat_totals[cat]
        target_pct = budgets.get(cat, Decimal(0))
        target_rp = float(target_pct) * float(r["income"])
        ws.cell(row=cur, column=2, value=cat_label.get(cat, cat)).font = BODY_FONT
        ws.cell(row=cur, column=3, value=float(target_pct)).number_format = PCT_FMT
        ws.cell(row=cur, column=4, value=target_rp).number_format = CURRENCY_FMT
        ws.cell(row=cur, column=5, value=_fmt(real)).number_format = CURRENCY_FMT
        ws.cell(row=cur, column=6, value=target_rp - _fmt(real)).number_format = CURRENCY_FMT
        for col in range(2, 7):
            cell = ws.cell(row=cur, column=col)
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col==2 else RIGHT
        cur += 1

    # Pie chart
    pie = PieChart()
    pie.title = "Distribusi Pengeluaran"
    data = Reference(ws, min_col=5, min_row=4, max_row=cur-1)
    cats = Reference(ws, min_col=2, min_row=5, max_row=cur-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList(showPercent=True)
    pie.height = 10; pie.width = 14
    ws.add_chart(pie, "H4")

    widths = {"A":2,"B":18,"C":12,"D":18,"E":18,"F":18}
    for c, w in widths.items(): ws.column_dimensions[c].width = w

    # Save
    wb.save(output_path)
    return output_path